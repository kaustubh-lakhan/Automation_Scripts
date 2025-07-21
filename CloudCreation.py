
#***************** MAKE SURE DATA IN EXCEL PROVIDED IS CORRECTLY SPELLED AND SELECTION DATA ALREADY PRESENT IN SYSTEM ********
# *******************************************************************************************************************
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support.select import Select

# importing excel and excelsheet
import openpyxl
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

##please enter the file path for the excel sheet
##DO NOT REMOVE r"" from the path
file = r"filePath/data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet =workbook.active

#configuration code / waiting code.
ops = webdriver.EdgeOptions()
ops.add_argument('--disable-notifications')
##add driver path for respective browser to your folder structure
service_object = Service(r"filePath\msedgedriver.exe")
driver = webdriver.Edge(service=service_object, options=ops)
driver.implicitly_wait(400)

#application URL
driver.get("https://org.tenrox.net/")

#Initial work through script
driver.find_element(By.ID, "menuImg55").click()
driver.switch_to.frame("fr55")
driver.find_element(By.ID, "lblUSERSSub8").click()


i=4 # "i" maintaines iframe number
c=1 # "c" maintaines column number

for j in range(7,9):
#as per no on excel sheet put range here here it will go from 2 to 10 . 10 is not inclusive
#so it will go from 2 to 9
    c=1
    driver.switch_to.default_content()
    driver.switch_to.frame("t4")  # 4

    #click on "create user" button
    driver.find_element(By.ID, "btnCreate").click()

    driver.switch_to.default_content()
    i += 1
    driver.switch_to.frame("t" + str(i))  # 5

    #last Name
    driver.find_element(By.ID, "txtLastName").send_keys(sheet.cell(j,c).value)
    c += 1

    # input("Do you want to proceed ?")
    #first Name
    driver.find_element(By.ID, "txtFirstName").send_keys(sheet.cell(j,c).value)
    c += 1
    # input("do you want to finish ?")
    print(sheet.cell(j,c).value)

    #ID
    driver.find_element(By.ID, "txtId").send_keys(sheet.cell(j,c).value)
    c += 1
    # input("do you want to finish ?")
    #Email
    driver.find_element(By.ID, "txtEmail").send_keys(sheet.cell(j,c).value)
    c += 1
    # input("do you want to finish ?")


    #Master Site
    driver.find_element(By.ID, "brwMasterSiteB").click()
    driver.switch_to.default_content()
    i += 1
    driver.switch_to.frame("t" + str(i))  # 6
    driver.find_element(By.ID, "searchForInputLeft").send_keys(sheet.cell(j,c).value)
    c += 1
    driver.find_element(By.ID, "searchBtnLeft").click()
    driver.find_element(By.XPATH, "//table[@id='tableBodyLeft0']//tr/td[1]").click()
    # input("do you want to finish ?")

    # for active site
    driver.switch_to.default_content()
    i -= 1
    driver.switch_to.frame("t" + str(i))  # 5
    driver.find_element(By.ID, "brwActiveSiteB").click()
    driver.switch_to.default_content()
    i += 2
    driver.switch_to.frame("t" + str(i))  # 7
    driver.find_element(By.ID, "searchForInputLeft").send_keys(sheet.cell(j,c).value)
    c += 1
    driver.find_element(By.ID, "searchBtnLeft").click()
    driver.find_element(By.XPATH, "//table[@id='tableBodyLeft0']//tr/td[1]").click()

    # for title
    i -= 2
    driver.switch_to.default_content()
    driver.switch_to.frame("t" + str(i))  # 5
    driver.find_element(By.ID, "brwTitleB").click()
    driver.switch_to.default_content()
    i += 3
    driver.switch_to.frame("t" + str(i))  # 8
    driver.find_element(By.ID, "searchForInputLeft").send_keys(sheet.cell(j,c).value)
    driver.find_element(By.ID, "searchBtnLeft").click()
    titles = driver.find_elements(By.XPATH, "//tbody//tr//a")
    for title in titles:
        if str(title.text) == str(sheet.cell(j,c).value):
            print(sheet.cell(j,c).value)
            title.click()
            break

    # approval group
    driver.switch_to.default_content()
    i -= 3
    c += 1
    driver.switch_to.frame("t" + str(i))  # 5
    driver.find_element(By.ID, "brwApprovalGroupB").click()
    driver.switch_to.default_content()
    i += 4
    driver.switch_to.frame("t" + str(i))  # 9
    driver.find_element(By.ID, "searchForInputLeft").send_keys(sheet.cell(j,c).value)
    c += 1
    driver.find_element(By.ID, "searchBtnLeft").click()
    driver.find_element(By.XPATH, "//table[@id='tableBodyLeft0']//tr/td[1]").click()

    # functional group
    driver.switch_to.default_content()
    i -= 4
    driver.switch_to.frame("t" + str(i))  # 5
    driver.find_element(By.ID, "brwFunctionalGroupB").click()
    driver.switch_to.default_content()
    i += 5
    driver.switch_to.frame("t" + str(i))  # 10
    driver.find_element(By.ID, "searchForInputLeft").send_keys(sheet.cell(j,c).value)
    c += 1
    driver.find_element(By.ID, "searchBtnLeft").click()
    driver.find_element(By.XPATH, "//table[@id='tableBodyLeft0']//tr/td[1]").click()

    # Resource group
    driver.switch_to.default_content()
    i -= 5
    driver.switch_to.frame("t" + str(i))  # 5
    driver.find_element(By.ID, "brwResourceGroupB").click()
    driver.switch_to.default_content()
    i += 6
    driver.switch_to.frame("t" + str(i))  # 11
    driver.find_element(By.ID, "searchForInputLeft").send_keys(sheet.cell(j,c).value)
    c += 1
    driver.find_element(By.ID, "searchBtnLeft").click()
    driver.find_element(By.XPATH, "//table[@id='tableBodyLeft0']//tr/td[1]").click()

    # LogonName
    driver.switch_to.default_content()
    i -= 6
    driver.switch_to.frame("t" + str(i))  # 5
    driver.find_element(By.ID, "txtLogonName").clear()
    driver.find_element(By.ID, "txtLogonName").send_keys(sheet.cell(j,c).value)
    c += 1

    # Password
    driver.find_element(By.ID, "txtPassword").send_keys("welcome25")

    # confirm password
    driver.find_element(By.ID, "txtConfirmPassword").send_keys("welcome25")

    # security role
    driver.find_element(By.XPATH,
                        "//*[@id='cmbSecProfileText_sCombo']/div/span").click()
    userAccount = driver.find_elements(By.XPATH, "//div[@class='cmbPopUpWrapper']/div/a")
    for choice in userAccount:
        if choice.text == sheet.cell(j,c).value:
            choice.click()
            if choice.text == "01-Standard User":
                driver.find_element(By.XPATH,"//*[@id='cmbSecProfileText_sCombo']/div/span").click()
            break;

    c += 1
    # input("do you want to put hire date ?")

    # hire date  - check the date/month/year - if need go thourgh service date of its dynamic date picker
    driver.find_element(By.XPATH, "//button[@name='Cal_ctrlHD']//i[@class='icomoon-calendar']").click()
    Month_picker = Select(driver.find_element(By.XPATH, "//select[@class='ui-datepicker-month']"))
    Month_picker.select_by_visible_text("April")
    Month_picker = Select(driver.find_element(By.XPATH, "//select[@class='ui-datepicker-year']"))
    Month_picker.select_by_visible_text("2025")
    allDates = driver.find_elements(By.XPATH, "//table[@class='ui-datepicker-calendar']//tbody/tr/td/div//a")
    for date in allDates:
        if date.text == "7":
            date.click()
            break

    #service date selection
    service_date = driver.find_element(By.XPATH, "//button[@name='Cal_ctrlSD']//i[@class='icomoon-calendar']")
    driver.execute_script("arguments[0].click();", service_date)

    # open ended
    # driver.find_element(By.XPATH, "//span[normalize-space()='Open Ended']").click()
    # time.sleep(2)
    # # driver.find_element(By.XPATH, "//span[normalize-space()='Open Ended']").click()
    # c+=1
    #
    # print("set date open ended")

    # date picker - *********** DO NOT DELETE ************
    # mMAKE SURE YOUR EXCEL DATA COLUMN HAVE YEAR TO TO SELECTED
    Month_picker = Select(driver.find_element(By.XPATH, "//select[@class='ui-datepicker-month']"))
    Month_picker.select_by_visible_text("January")
    if(sheet.cell(j, c).value<2015):
        Year_picker = Select(driver.find_element(By.XPATH, "//select[@class='ui-datepicker-year']"))
        Year_picker.select_by_visible_text("2015")
        if (sheet.cell(j, c).value < 2005):
            Year_picker = Select(driver.find_element(By.XPATH, "//select[@class='ui-datepicker-year']"))
            Year_picker.select_by_visible_text("2005")
            # Year_picker = Select(driver.find_element(By.XPATH, "//select[@class='ui-datepicker-year']"))
            # Year_picker.select_by_visible_text(str(sheet.cell(j, c).value))
            if (sheet.cell(j, c).value < 1995):
                Year_picker = Select(driver.find_element(By.XPATH, "//select[@class='ui-datepicker-year']"))
                Year_picker.select_by_visible_text("1995")
                if(sheet.cell(j, c).value < 1985):
                    Year_picker = Select(driver.find_element(By.XPATH, "//select[@class='ui-datepicker-year']"))
                    Year_picker.select_by_visible_text("1985")
                    Year_picker = Select(driver.find_element(By.XPATH, "//select[@class='ui-datepicker-year']"))
                    Year_picker.select_by_visible_text(str(sheet.cell(j, c).value))
                else:
                    Year_picker = Select(driver.find_element(By.XPATH, "//select[@class='ui-datepicker-year']"))
                    Year_picker.select_by_visible_text(str(sheet.cell(j, c).value))
            else:
                Year_picker = Select(driver.find_element(By.XPATH, "//select[@class='ui-datepicker-year']"))
                Year_picker.select_by_visible_text(str(sheet.cell(j, c).value))
        else:
            Year_picker = Select(driver.find_element(By.XPATH, "//select[@class='ui-datepicker-year']"))
            Year_picker.select_by_visible_text(str(sheet.cell(j, c).value))
    else :
        Year_picker = Select(driver.find_element(By.XPATH, "//select[@class='ui-datepicker-year']"))
        Year_picker.select_by_visible_text(str(sheet.cell(j,c).value))
    c+=1
    allDates = driver.find_elements(By.XPATH, "//table[@class='ui-datepicker-calendar']//tbody/tr/td/div//a")
    for date in allDates:
        if date.text == "1":
            date.click()
            break

    # user type
    driver.find_element(By.XPATH,"//*[@id='cmbUT_sCombo']/div/span").click()
    userAccount = driver.find_elements(By.XPATH, "//div[@class='cmbPopUpWrapper']/div/a")
    print(userAccount)
    for choice in userAccount:
        if choice.text == sheet.cell(j,c).value:
            choice.click()
            if choice.text == "Employee-Exempted":
                driver.find_element(By.XPATH,"//*[@id='cmbUT_sCombo']/div/span").click()
            break;
    c+=1

    print("proceeding with holiday set")
    time.sleep(2)
    # holiday set
    driver.find_element(By.XPATH, "//div[@id='brwHolidaySetB']").click()
    driver.switch_to.default_content()
    i += 7
    driver.switch_to.frame("t" + str(i))  # 12
#   #############################################
    driver.find_element(By.ID, "searchForInputLeft").send_keys(sheet.cell(j,c).value)
    c += 1
    driver.find_element(By.ID, "searchBtnLeft").click()
    driver.find_element(By.XPATH, "//table[@id='tableBodyLeft0']//tr/td[1]").click()
    ###########################################
    # set to none
    # driver.find_element(By.XPATH, "//button[@id='btnNone']").click();
    # print("inside the holiday set, set to none")
    # c += 1;
    # ############################################
    driver.switch_to.default_content()
    i -= 7
    driver.switch_to.frame("t" + str(i))  # 5

    # input("resource type?")
    # resource Type
    driver.find_element(By.ID, "brwResourceTypeB").click()
    driver.switch_to.default_content()
    i += 8
    driver.switch_to.frame("t" + str(i))  # 13
    driver.find_element(By.ID, "searchForInputLeft").send_keys(sheet.cell(j,c).value)
    c+=1
    driver.find_element(By.ID, "searchBtnLeft").click()
    driver.find_element(By.XPATH, "//table[@id='tableBodyLeft0']//tr/td[1]").click()
    driver.switch_to.default_content()
    i -= 8
    driver.switch_to.frame("t" + str(i))  # 5

    # Company
    driver.find_element(By.ID, "brwCompanyB").click()
    driver.switch_to.default_content()
    i += 9
    driver.switch_to.frame("t" + str(i))  # 14
    driver.find_element(By.ID, "searchForInputLeft").send_keys(sheet.cell(j,c).value)
    c+=1
    driver.find_element(By.ID, "searchBtnLeft").click()
    driver.find_element(By.XPATH, "//table[@id='tableBodyLeft0']//tr/td[1]").click()
    driver.switch_to.default_content()
    i -= 9
    driver.switch_to.frame("t" + str(i))  # 5

    #moving to Eaton fields
    driver.find_element(By.XPATH, "//*[@id='li4']/a/span").click()


    #Flexible time schedule
    driver.find_element(By.XPATH, "//*[@id='7_1_202_1']//span[@class='cmbArrow cmbTextInputArrow']").click()
    FlexibleSchedules = driver.find_elements(By.XPATH, "//div[@class='cmbPopUpWrapper']/div/a")
    for schedule in FlexibleSchedules:
        if schedule.text == sheet.cell(j,c).value:
            schedule.click()
            if schedule.text == "Standard Work Week":
                driver.find_element(By.XPATH, "//*[@id='7_1_202_1']//span[@class='cmbArrow cmbTextInputArrow']").click()
                time.sleep(1)
            break;
    c+=1

    # Active user account
    driver.find_element(By.XPATH, "//*[@id='7_1_215_1']//span[@class='cmbArrow cmbTextInputArrow']").click()
    FlexibleSchedules = driver.find_elements(By.XPATH, "//div[@class='cmbPopUpWrapper']/div/a")
    for schedule in FlexibleSchedules:
        if schedule.text == "Yes":
            schedule.click()
            # driver.find_element(By.XPATH, "//*[@id='7_1_215_1']//span[@class='cmbArrow cmbTextInputArrow']").click()
            # schedule.click()
            time.sleep(1)
            break;


    # status
    driver.find_element(By.XPATH, "//*[@id='7_1_216_1']//span[@class='cmbArrow cmbTextInputArrow']").click()
    FlexibleSchedules = driver.find_elements(By.XPATH, "//div[@class='cmbPopUpWrapper']/div/a")
    for schedule in FlexibleSchedules:
        if schedule.text == "Employee":
            schedule.click()
            driver.find_element(By.XPATH, "//*[@id='7_1_216_1']//span[@class='cmbArrow cmbTextInputArrow']").click()
            time.sleep(1)
            break;

    # save
    # input("do you want wait ad procees ?")
    driver.find_element(By.XPATH, "//*[@id='btnStandaloneSave']").click()
    # input("do you want to proceed for scope ?")

    # input("do you want to proceed ?")
    #scope
    driver.switch_to.default_content()
    driver.switch_to.frame("t" + str(i))
    driver.find_element(By.XPATH,
                        "//div[@id='leftScope']//span[@class='text-label ellipsis ignoreTextTransform tenroxTheme2'][normalize-space()='Scope']").click()
    time.sleep(3)
    iframeScope = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//iframe[contains(@id,'frm_pvScope')]")))
    driver.switch_to.frame(iframeScope)
    time.sleep(2);
    driver.find_element(By.XPATH, "//input[@id='searchForInputLeft']").clear()
    driver.find_element(By.XPATH, "//input[@id='searchForInputLeft']").send_keys(sheet.cell(j,5).value)
    print(sheet.cell(j,5).value)
    driver.find_element(By.XPATH, "//div[@id='searchBtnLeft']//div[@class='icon-icomoon-new-search t-grid-search-icon tenroxTheme2']").click()
    time.sleep(4)
    rows = driver.find_elements(By.XPATH,"//table[@id='tableBodyLeft0']//tr")
    for row in rows:
        name = row.find_element(By.XPATH,".//a")
        if name.text == sheet.cell(j,5).value :
            row.find_element(By.XPATH,"//div[@class='squaredCheckbox']").click()
            print("proceeding with selecting" + sheet.cell(j,5).value)
            break;
    driver.find_element(By.XPATH,
                        "//body[1]/form[1]/div[3]/section[1]/div[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/div[2]/table[1]/tbody[1]/tr[1]/td[1]").click()
    time.sleep(2)
    driver.find_element(By.XPATH, "//input[@id='searchForInputLeft']").clear()
    # driver.find_element(By.XPATH, "//input[@id='searchForInputLeft']").send_keys(sheet.cell(j,6).value)
    # time.sleep(3)
    # driver.find_element(By.XPATH,
    #                     "//div[@id='searchBtnLeft']//div[@class='icon-icomoon-new-search t-grid-search-icon tenroxTheme2']").click()
    # driver.find_element(By.XPATH, "//label[@for='chk_item_Left386']").click()  ##need to change every time
    driver.find_element(By.XPATH, "//input[@id='searchForInputLeft']").send_keys(sheet.cell(j,6).value)
    print(sheet.cell(j,6).value)
    driver.find_element(By.XPATH, "//div[@id='searchBtnLeft']//div[@class='icon-icomoon-new-search t-grid-search-icon tenroxTheme2']").click()
    time.sleep(4)
    rows = driver.find_elements(By.XPATH,"//table[@id='tableBodyLeft0']//tr")
    for row in rows:
        print(row)
        name = row.find_element(By.XPATH,".//a")
        print(name.text)
        if name.text == sheet.cell(j,6).value :
            row.find_element(By.XPATH,"//div[@class='squaredCheckbox']").click()
            print("proceeding with selecting" + sheet.cell(j,6).value)
            break;
    driver.find_element(By.XPATH,
                        "//body[1]/form[1]/div[3]/section[1]/div[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/div[2]/table[1]/tbody[1]/tr[1]/td[1]").click()
    # input("do you want to proceed ?")
    driver.find_element(By.XPATH,
                        "//body[1]/form[1]/div[3]/section[1]/div[1]/table[1]/tbody[1]/tr[2]/td[1]/div[1]/div[2]/table[1]/tbody[1]/tr[1]/td[1]").click()
    # input("do you want to proceed ?")

    # cost center - MAKE SURE ALL COST CENTER PRESENT IN THE SYSTEM FIRST BEFR=ORE YOU RUN SCRIPT OTHER WISE IT WILL GIVE ERROR
    driver.switch_to.default_content()
    driver.switch_to.frame("t" + str(i))  # 5
    driver.find_element(By.XPATH, "//span[normalize-space()='Business_units']").click()
    iframeCostC = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//iframe[contains(@id,'frm_pvBusinessUnits')]")))
    driver.switch_to.frame(iframeCostC)
    driver.find_element(By.XPATH, "//button[@id='btnNew_pageGrid']").click()
    driver.find_element(By.XPATH, "//div[@id='bpageGrid_Name__newB']").click()
    driver.switch_to.default_content()
    i += 10
    driver.switch_to.frame("t" + str(i))  # 15
    driver.find_element(By.XPATH, "//input[@id='searchForInputLeft']").send_keys(sheet.cell(j,c).value)
    # c += 1
    driver.find_element(By.XPATH, "//div[@class='icon-icomoon-new-search t-grid-search-icon tenroxTheme2']").click()
    # input("do you want to proceed ?")
    driver.find_element(By.XPATH, "//table[@id='tableBodyLeft0']//tbody/tr/td/div").click()
    driver.switch_to.default_content()
    i -= 10
    driver.switch_to.frame("t" + str(i))  # 5
    driver.switch_to.frame(iframeCostC)
    driver.find_element(By.XPATH, "//div[@id='save_pageGrid_new']").click()

    # notes
    driver.switch_to.default_content()
    driver.switch_to.frame("t" + str(i))  # 5
    driver.find_element(By.XPATH, "//span[normalize-space()='Attachments']").click()
    iframeAttachment= WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH,"//iframe[contains(@id,'frm_pvAttachments')]")))
    driver.switch_to.frame(iframeAttachment)
    # clicking on new
    driver.find_element(By.XPATH, "//button[@id='btnNew_NoteGrid']").click()
    # switching to window for notes
    driver.switch_to.default_content()
    i += 11
    driver.switch_to.frame("t" + str(i))  # 16
    driver.find_element(By.XPATH, "//input[@id='cmbNoteType']").click()
    # drop down
    driver.find_element(By.XPATH, "//a[normalize-space()='Notice']").click()
    # PROVIDE NOTE IN BELOW SEND KEYS PART / WE CAN DYNAMICALLY GET FROM EXCEL NEED TO WRITE CODE FOR IT
    driver.find_element(By.XPATH, "//textarea[@id='txtNoteDesc']").send_keys(
        "RITM1216497 - Account has been created - KL")
    driver.find_element(By.XPATH, "//button[@id='btnSave']").click()
    driver.switch_to.default_content()
    i -= 11
    driver.switch_to.frame("t" + str(i)) #5
    driver.find_element(By.XPATH, "//button[@id='btnStandaloneCancel']").click()
    i += 11



#RUN THE SCRIPT SOMETIME IT WILL END WITHOUT ANY ERROR - PLEASE RELAUNCH THE SCRIPT
#***************** MAKE SURE DATA IN EXCEL PROVIDED IS CORRECTLY SPELLED AND ALREADY PRESENT IN SYSTEM ********
#DEBUGG IT AS PER NEED BY ADDING INPUT/assert STATEMENT . USE THEM AS A BREAK FOR SCRIPT.